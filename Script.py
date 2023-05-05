import openpyxl

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Add headers to the worksheet
worksheet['A1'] = 'Name'
worksheet['B1'] = 'Age'
worksheet['C1'] = 'City'

# Prompt the user for input and write to the worksheet
row = 2
while True:
    name = input("Enter a name (or 'q' to quit): ")
    if name == 'q':
        break
    age = input("Enter an age: ")
    city = input("Enter a city: ")

    worksheet.cell(row=row, column=1, value=name)
    worksheet.cell(row=row, column=2, value=age)
    worksheet.cell(row=row, column=3, value=city)
    
    row += 1

# Save the workbook
workbook.save('Data.xlsx')
